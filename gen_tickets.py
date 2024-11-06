from openpyxl import load_workbook
from datetime import datetime, timedelta
import random

# Define lists of sample names and email domains
names = ["Alice", "Bob", "Charlie", "Diana", "Evan", "Fiona", "George", "Hannah", "Ivan", "Jasmine"]
domains = ["gmail.com", "hotmail.com", "outlook.com", "yahoo.com", "protonmail.com"]

# Generate sample data with conditional Date Resolved and randomized names/emails
def generate_ticket_data(num_tickets):
    data = []
    for i in range(1, num_tickets + 1):
        name = random.choice(names)
        domain = random.choice(domains)
        status = random.choice(["Open", "In Progress", "Resolved", "Closed"])
        n = random.randint(0, 100)
        
        ticket = {
            "Ticket ID": f"TKT{str(i).zfill(4)}",
            "Date Submitted": (datetime.now() - timedelta(days=n)).strftime("%Y-%m-%d %H:%M:%S"),
            "Submitted By": f"{name.lower()}{n}@{domain}",
            "Issue Description": "Example issue",
            "Priority": random.choice(["Low", "Medium", "High"]),
            "Status": status,
            "Assigned To": f"{name.lower()}{n}@{domain}",
            "Date Resolved": (datetime.now().strftime("%Y-%m-%d %H:%M:%S") if status == "Resolved" else None),
        }
        data.append(ticket)
    return data

# Load the existing Excel workbook and select the active sheet
excel_file = "Ticket_Log.xlsx"
book = load_workbook(excel_file)
sheet = book.active

# Determine the starting row for appending data
last_row = sheet.max_row + 1

# Generate data and append each row to the sheet
ticket_data = generate_ticket_data(100)  # Generate 100 tickets
for ticket in ticket_data:
    # Write each column of the ticket as a new row in the sheet
    sheet.cell(row=last_row, column=1, value=ticket["Ticket ID"])
    sheet.cell(row=last_row, column=2, value=ticket["Date Submitted"])
    sheet.cell(row=last_row, column=3, value=ticket["Submitted By"])
    sheet.cell(row=last_row, column=4, value=ticket["Issue Description"])
    sheet.cell(row=last_row, column=5, value=ticket["Priority"])
    sheet.cell(row=last_row, column=6, value=ticket["Status"])
    sheet.cell(row=last_row, column=7, value=ticket["Assigned To"])
    sheet.cell(row=last_row, column=8, value=ticket["Date Resolved"])
    last_row += 1  # Move to the next row for each ticket

# Save the workbook after appending
book.save(excel_file)
book.close()
